# importing openpyxl module
from boto3 import client
from openpyxl import Workbook, load_workbook

# Give the document numbers you want to compare in the following Excel.
path = "likp1007.XLSX"

sourceDirectory = 'input/valid/'
destinationDirectory = 'codeTrigger/'
Bucket='opsandsrvosv-asnedi-prod-us-west-2'


# workbook object is created
wb_obj = load_workbook(path)

wb = Workbook()

ws = wb.create_sheet(title="foundShipments")
ws.append(["Key"])

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

print(m_row)


#Check if it is valid.
valids = []
#Count the number of records copied
count = 0
page_number = 0
#Avoid duplicates
checked = {}

print(count)

# Loop will append all values
# of first column
for i in range(1, m_row + 1):
	cell_obj = sheet_obj.cell(row = i, column = 1)
	valids.append(cell_obj.value)
    

print(valids)
#print(len(valids))


s3 = client('s3')

paginator = s3.get_paginator('list_objects_v2')

#Get all the pages from the bucket
pages = paginator.paginate(Bucket=Bucket, Prefix=sourceDirectory)

for page in pages:
    page_number += 1
    print(page_number)

    #For each document in the page
    for obj in page['Contents']:

        #Extracting the Document number from the filepath 
        Key = obj['Key'].split('/')
        item_ = Key[-1].split('-')

        #Ignoring duplicates
        # if item_[0] in checked:
        #     continue
        # else:
            #Adding the document number to the list for duplicate check
            # checked[item_[0]] = True

             #If mising according to the list provided
        if item_[0] in valids:
            ws.append([item_[0]])
            count += 1

            copy_source = Bucket + '/' + obj['Key']
            copyDestination = destinationDirectory + obj['Key']

            #aws cli command to copy object
            response = s3.copy_object(CopySource = copy_source, Bucket=Bucket, Key=copyDestination)  
            #print(count, copy_source + ' - ' + Bucket + '/' + copyDestination)
            print(count, response)


        else:
            continue

#Printing total count after the run is complete
print(count)
wb.save('copied_valids.xlsx')