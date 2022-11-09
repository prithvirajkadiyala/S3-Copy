from boto3 import client
from openpyxl import Workbook, load_workbook

path = "filesxlsx/policies.xlsx"
count = 0
# workbook object is created
wb_obj = load_workbook(path)
 
sheet_obj = wb_obj.active
 
#print total number of column
m_row = sheet_obj.max_row
print(m_row)

iam = client('iam')
    
tags=[
        {
            "Key": "nike-department",
            "Value": "global order and logistics"
        },
        {
            "Key": "nike-owner",
            "Value": "scott.marien@nike.com"
        },
        {
            "Key": "nike-distributionlist",
            "Value": "lst-sec.opsandservices.integration@nike.com"
        },
        {
            "Key": "nike-domain",
            "Value": "order status visibility"
        },
        {
            "Key": "nike-application",
            "Value": "opsandsrvosv-test"
        },
        {
            "Key": "nike-environment",
            "Value": "test"
        }
    ]

for obj in range(2, m_row + 1):
    count += 1
    cell_obj = sheet_obj.cell(row = obj, column = 1)
    print()
    response = iam.tag_policy(
        PolicyArn=cell_obj.value,
        Tags=tags
    )       
    print(count, cell_obj.value, response)